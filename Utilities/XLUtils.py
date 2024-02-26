import openpyxl


def getRowCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row


def getColumnCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_column


def readData(file, sheetname, rownum, colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=rownum, column=colnum).value


def readExpData(file, sheetname, function_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    # Find the row number where the function name matches
    for row in sheet.iter_rows(values_only=True):
        if row[0] == function_name:
            return row[2]  # Row number of Test Data field
    return None  # Function name not found in the sheet


def writeData(file, sheetname, rownum, colnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rownum, column=colnum).value = data
    workbook.save(file)
